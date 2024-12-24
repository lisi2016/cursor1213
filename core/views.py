# -*- coding: utf-8 -*-

from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponse, FileResponse, StreamingHttpResponse
from django.core.exceptions import PermissionDenied
import pandas as pd
from datetime import datetime
from .models import User, Assignment, DistributionLog
from .machine_monitor import MachineMonitor
from .assignment_handler import AssignmentDistributor
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.contrib import messages
import os
import urllib.parse
from django.core.files.base import ContentFile
from django.conf import settings
import shutil
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
import json
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.encoding import force_str
from django.contrib.sessions.models import Session
from django.db import transaction
import logging

logger = logging.getLogger(__name__)

@require_http_methods(["GET", "POST"])
def login_view(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            if not user.is_teacher:
                try:
                    # 获取用户IP
                    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                    ip = x_forwarded_for.split(',')[0].strip() if x_forwarded_for else request.META.get('REMOTE_ADDR')
                    
                    # 更新用户信息
                    user.ip_address = ip
                    user.last_login = timezone.now()
                    user.save(update_fields=['ip_address', 'last_login'])
                    
                    # 准备 WebSocket 消息
                    message = {
                        "type": "student_status",
                        "data": {
                            "action": "login",
                            "student_id": user.student_id,
                            "name": user.get_full_name() or user.username,
                            "ip": ip,
                            "last_login": timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')
                        }
                    }
                    
                    print("="*50)
                    print(f"学生登录信息:")
                    print(f"用户名: {username}")
                    print(f"学号: {user.student_id}")
                    print(f"IP地址: {ip}")
                    print(f"登录时间: {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}")
                    
                    # 发送 WebSocket 消息
                    try:
                        channel_layer = get_channel_layer()
                        if not channel_layer:
                            print("错误: 无法获取 channel layer")
                            return redirect('student_dashboard')
                            
                        # 检查教师组是否存在
                        groups = getattr(channel_layer, 'groups', {})
                        print(f"当前活动的 channel layer 组: {groups}")
                        print(f"准备发送的消息: {message}")
                        
                        # 发送消息到教师组
                        async_to_sync(channel_layer.group_send)(
                            "teacher_group",
                            {
                                "type": "student_status",
                                "data": message['data']
                            }
                        )
                        print("WebSocket 消息发送成功")
                        print("="*50)
                        
                    except Exception as e:
                        print(f"发送 WebSocket 消息时出错: {str(e)}")
                        print(f"错误类型: {type(e)}")
                        print(f"错误详情: {str(e)}")
                        
                    return redirect('student_dashboard')
                    
                except Exception as e:
                    print(f"处理学生登录时出错: {str(e)}")
            
            return redirect('teacher_dashboard' if user.is_teacher else 'student_dashboard')
        else:
            logger.warning(f"Login failed for username: {username}")
            
    return render(request, 'core/login.html')

@login_required
def teacher_dashboard(request):
    if not request.user.is_teacher:
        raise PermissionDenied
        
    # 获取所有作业任务
    assignments = Assignment.objects.all().select_related('assigned_to').order_by('-upload_time')
    
    # 获取所有班级列表（用于导出成绩的班级筛选）
    class_list = User.objects.filter(
        is_teacher=False
    ).values_list('class_name', flat=True).distinct().order_by('class_name')
    
    # 获取在线学生
    students = User.objects.filter(
        is_teacher=False,
        last_login__isnull=False,
        last_login__gte=timezone.now() - timezone.timedelta(minutes=5)
    )
    
    context = {
        'assignments': assignments,
        'class_list': class_list,  # 加班级列表到上下文
        'students': students,
    }
    return render(request, 'core/teacher_dashboard.html', context)

@login_required
def student_dashboard(request):
    if request.user.is_teacher:
        raise PermissionDenied
        
    assignments = Assignment.objects.filter(assigned_to=request.user)
    context = {
        'assignments': assignments
    }
    return render(request, 'core/student_dashboard.html', context)

@login_required
def import_students(request):
    if not request.user.is_teacher:
        raise PermissionDenied
        
    if request.method == 'POST' and request.FILES.get('student_file'):
        try:
            df = pd.read_excel(request.FILES['student_file'])
            
            # 验证必需的列是否存在
            required_columns = ['学号', '姓名', '班级']
            if not all(col in df.columns for col in required_columns):
                messages.error(request, '文件格式错误，请使用正确的导入模板')
                return redirect('import_students')
            
            # 验证数据有效性
            success_count = 0
            error_messages = []
            
            for index, row in df.iterrows():
                try:
                    # 验证学号格式
                    if not str(row['学号']).isdigit():
                        error_messages.append(f"第{index+2}行：学号必须为数字")
                        continue
                        
                    # 验证姓名不为空
                    if pd.isna(row['姓名']) or str(row['姓名']).strip() == '':
                        error_messages.append(f"第{index+2}行姓名不能为空")
                        continue
                    
                    # 查学号是否已存在
                    if User.objects.filter(username=str(row['学号'])).exists():
                        error_messages.append(f"第{index+2}行：学号{row['学号']}已存在")
                        continue
                        
                    # 创建用户
                    User.objects.create_user(
                        username=str(row['学号']),
                        password=str(row['学号']),  # 初始密码设为学号
                        student_id=str(row['学号']),
                        first_name=str(row['姓名']),
                        class_name=str(row['班级']),
                        is_teacher=False
                    )
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f"第{index+2}行：{str(e)}")
            
            # 导入结果
            if success_count > 0:
                messages.success(request, f'成功导入{success_count}条记录')
            if error_messages:
                messages.warning(request, '\n'.join(error_messages))
            
            if success_count > 0 and not error_messages:
                return redirect('teacher_dashboard')
            
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
    
    return render(request, 'core/import_students.html')

@login_required
@ensure_csrf_cookie
def distribute_assignments(request):
    print("\n=== Request Debug Info ===")
    print(f"Method: {request.method}")
    print(f"Path: {request.path}")
    print(f"User: {request.user.username} (Teacher: {request.user.is_teacher})")
    print("\nHeaders:")
    for key, value in request.headers.items():
        print(f"{key}: {value}")
    
    print("\nPOST data:")
    for key, value in request.POST.items():
        print(f"{key}: {value}")
    
    print("\nFiles:")
    for key, value in request.FILES.items():
        print(f"{key}: {value.name} ({value.size} bytes)")
    
    print(f"\nCSRF Token from header: {request.META.get('HTTP_X_CSRFTOKEN')}")
    print(f"Session ID: {request.session.session_key}")
    print(f"Cookie Header: {request.META.get('HTTP_COOKIE')}")
    
    if not request.user.is_teacher:
        print("错误: 非教师用户")
        return JsonResponse({
            'status': 'error',
            'message': '权限不足'
        }, status=403)
    
    try:
        # 1. 检查文件
        files = request.FILES.getlist('assignments')
        print(f"接收到的文件: {[f.name for f in files]}")
        
        if not files:
            print("错误: 没有文件")
            return JsonResponse({
                'status': 'error',
                'message': '请选择要分发的文件'
            }, status=400)

        # 2. 检查学生ID列表
        student_ids_json = request.POST.get('student_ids')
        print(f"接收到的学生ID列表: {student_ids_json}")
        
        if not student_ids_json:
            print("错误: 没有学生ID列表")
            return JsonResponse({
                'status': 'error',
                'message': '没有选择学生'
            }, status=400)
            
        try:
            student_ids = json.loads(student_ids_json)
            print(f"解析后的学生ID列表: {student_ids}")
        except json.JSONDecodeError as e:
            print(f"错误: 无效的学生ID列表格式 - {e}")
            return JsonResponse({
                'status': 'error',
                'message': '无效的学生ID列表格式'
            }, status=400)

        # 3. 获取在线学生
        online_students = list(User.objects.filter(
            student_id__in=student_ids,
            is_teacher=False,
            last_login__isnull=False,
            last_login__gte=timezone.now() - timezone.timedelta(minutes=5)
        ).order_by('student_id'))
        
        print(f"查询到的在线学生: {[f'{s.student_id}({s.first_name})' for s in online_students]}")

        if not online_students:
            print("错误: 没有找到在线学生")
            return JsonResponse({
                'status': 'error',
                'message': '所选学生已不在线'
            }, status=400)

        # 4. 分发作业
        success_count = 0
        error_messages = []
        
        try:
            with transaction.atomic():
                for file in files:
                    student_index = success_count % len(online_students)
                    student = online_students[student_index]
                    
                    try:
                        print(f"正在分发文件 {file.name} 给学生 {student.student_id}")
                        
                        assignment = Assignment.objects.create(
                            file_name=file.name,
                            assigned_to=student,
                            upload_time=timezone.now()
                        )
                        
                        file_name = f"{student.student_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}_{file.name}"
                        file.seek(0)
                        assignment.file_path.save(file_name, ContentFile(file.read()), save=True)
                        
                        print(f"成功分发文件 {file.name} 给学生 {student.student_id}")
                        success_count += 1
                        
                    except Exception as e:
                        error_msg = f'分配给学生 {student.first_name}({student.student_id}) 失败: {str(e)}'
                        print(f"错误: {error_msg}")
                        error_messages.append(error_msg)
        except Exception as e:
            print(f"错误: 事务执行失败 - {e}")
            return JsonResponse({
                'status': 'error',
                'message': f'数据库事务失败: {str(e)}'
            }, status=500)

        print(f"分发完成: 成功 {success_count} 个, 失败 {len(error_messages)} 个")
        if success_count > 0:
            message = f'成功分发 {success_count} 个作业'
            if error_messages:
                message += f'，但有 {len(error_messages)} 个分��失败'
            return JsonResponse({
                'status': 'success',
                'message': message
            })
        else:
            return JsonResponse({
                'status': 'error',
                'message': '分发失败：\n' + '\n'.join(error_messages)
            }, status=400)

    except Exception as e:
        print(f"错误: 未处理的异常 - {e}")
        import traceback
        print(f"错误堆栈: {traceback.format_exc()}")
        return JsonResponse({
            'status': 'error',
            'message': f'系统错误：{str(e)}'
        }, status=500)

@login_required
def download_assignment(request, assignment_id):
    try:
        assignment = Assignment.objects.get(id=assignment_id, assigned_to=request.user)
        
        print("开始处理下载请求:")
        print(f"- MEDIA_ROOT: {settings.MEDIA_ROOT}")
        print(f"- 文件相对路径: {assignment.file_path}")
        print(f"- 文件名: {assignment.file_name}")
        
        if not assignment.file_path:
            print("错误: 文件路径为空")
            return HttpResponse('文件不存在', status=404)
            
        try:
            file_path = os.path.join(settings.MEDIA_ROOT, str(assignment.file_path))
            print(f"- 完整文件路径: {file_path}")
            
            if not os.path.exists(file_path):
                print(f"错误: 文件不存在于路径: {file_path}")
                return HttpResponse('文件不存在', status=404)
                
            # 使用 StreamingHttpResponse 替代 FileResponse
            def file_iterator(file_path, chunk_size=8192):
                with open(file_path, 'rb') as f:
                    while True:
                        chunk = f.read(chunk_size)
                        if not chunk:
                            break
                        yield chunk

            response = StreamingHttpResponse(
                file_iterator(file_path),
                content_type='application/octet-stream'
            )
            
            # 设置文件名，支持中文
            encoded_filename = urllib.parse.quote(assignment.file_name)
            response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
            
            # 更新下载状态
            assignment.download_status = True
            assignment.download_time = timezone.now()
            assignment.save(update_fields=['download_status', 'download_time'])
            
            print(f"文件下载成功: {assignment.file_name}")
            return response
                
        except Exception as e:
            print(f"文件读取错误: {str(e)}")
            return HttpResponse('文件读取失败', status=500)
            
    except Assignment.DoesNotExist:
        print(f"作业不存在: {assignment_id}")
        return HttpResponse('作业不存在', status=404)
    except Exception as e:
        print(f"下载过程中发生错误: {str(e)}")
        return HttpResponse('下载失败', status=500)

@login_required
def grade_assignment(request, assignment_id):
    if request.method == 'POST':
        grade = request.POST.get('grade')
        
        # 验证成绩格式
        if not grade or grade not in ['A', 'B', 'C', 'D']:
            return JsonResponse({
                'status': 'error',
                'message': '无效的成绩等级'
            })
            
        try:
            assignment = Assignment.objects.get(id=assignment_id, assigned_to=request.user)
            
            # 检查是否已经评过分
            if assignment.grade:
                return JsonResponse({
                    'status': 'error',
                    'message': '该作业已经评过分'
                })
                
            # 更新成绩
            assignment.grade = grade
            assignment.save(update_fields=['grade'])
            
            return JsonResponse({
                'status': 'success',
                'grade_display': assignment.get_grade_display()
            })
            
        except Assignment.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': '作业不存在'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'评分失败：{str(e)}'
            })
            
    return JsonResponse({
        'status': 'error',
        'message': '方法不允许'
    })

@login_required
def export_grades(request):
    if not request.user.is_teacher:
        raise PermissionDenied
        
    try:
        # 获取选的班级
        class_name = request.GET.get('class_name')
        
        # 构建查询条件
        query = {}
        if class_name:
            query['assigned_to__class_name'] = class_name
            
        # 获取作业数据
        assignments = Assignment.objects.filter(
            **query
        ).select_related('assigned_to').order_by(
            'assigned_to__class_name',
            'assigned_to__student_id',
            '-upload_time'
        )
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩统计"
        
        # 设置表头
        headers = ['学号', '姓名', '班级', '作业名称', '成绩', '下载状态', '下载时间', '分发时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 写入数据
        for row, assignment in enumerate(assignments, 2):
            student = assignment.assigned_to
            ws.cell(row=row, column=1).value = student.student_id
            ws.cell(row=row, column=2).value = student.first_name
            ws.cell(row=row, column=3).value = student.class_name
            ws.cell(row=row, column=4).value = assignment.file_name
            ws.cell(row=row, column=5).value = assignment.get_grade_display() if assignment.grade else '未评分'
            ws.cell(row=row, column=6).value = '已下载' if assignment.download_status else '未下载'
            ws.cell(row=row, column=7).value = timezone.localtime(assignment.download_time).strftime('%Y-%m-%d %H:%M:%S') if assignment.download_time else '-'
            ws.cell(row=row, column=8).value = timezone.localtime(assignment.upload_time).strftime('%Y-%m-%d %H:%M:%S')
        
        # 设置列宽
        column_widths = [15, 15, 15, 40, 10, 10, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 设置文件名（包含班级息和时间戳）
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"成绩统计_{class_name or '全部'}_{timestamp}.xlsx"
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(filename)}'
        
        # 保存到响应
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"导出成绩时发生误: {str(e)}")
        messages.error(request, f'导出失败：{str(e)}')
        return redirect('teacher_dashboard')

# 添加一个新的视图来更新服务器状态
@login_required
def update_machine_status(request):
    # 如果是学生访问，返回空数据而不是抛出权限错误
    if not request.user.is_teacher:
        return JsonResponse({})
        
    # 获取最近5分钟内登录的所有学生
    five_minutes_ago = timezone.now() - timezone.timedelta(minutes=5)
    students = User.objects.filter(
        is_teacher=False,
        last_login__gte=five_minutes_ago,
        ip_address__isnull=False
    ).order_by('ip_address')  # 按IP地址排序
    
    machine_status = {}
    for student in students:
        local_login_time = timezone.localtime(student.last_login)
        machine_status[student.ip_address] = {
            'online': True,
            'student_name': f"{student.first_name}({student.student_id})",
            'last_login': local_login_time.strftime('%Y-%m-%d %H:%M:%S')
        }
    
    return JsonResponse(machine_status)

@login_required
def logout_view(request):
    logger.info(f"Processing logout for user: {request.user.username}")
    
    if not request.user.is_teacher:
        try:
            channel_layer = get_channel_layer()
            logger.info(f"Got channel layer for logout: {type(channel_layer)}")
            
            message = {
                "type": "student_status",
                "data": {
                    "action": "logout",
                    "student_id": request.user.student_id
                }
            }
            
            logger.info(f"Sending logout notification: {message}")
            
            async_to_sync(channel_layer.group_send)(
                "teacher_group",
                message
            )
            
            logger.info("Logout notification sent successfully")
            
            request.user.ip_address = None
            request.user.save(update_fields=['ip_address'])
            logger.info(f"Cleared IP address for user {request.user.username}")
            
        except Exception as e:
            logger.error(f"Error during logout notification: {str(e)}", exc_info=True)
    
    logout(request)
    logger.info(f"User {request.user.username} logged out successfully")
    return redirect('login')

def download_import_template(request):
    if not request.user.is_teacher:
        raise PermissionDenied
        
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息导入模板"
    
    # 设置表头
    headers = ['学号', '姓名', '班级']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        # 设置表头样式
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 示例数据
    example_data = [
        ['2024001', '张三', '计算机1班'],
        ['2024002', '李四', '算机1班']
    ]
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col).value = value
    
    # 设置宽度
    ws.column_dimensions['A'].width = 15  # 学号列
    ws.column_dimensions['B'].width = 15  # 姓名列
    ws.column_dimensions['C'].width = 20  # 班级列
    
    # 创建HTTP响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_import_template.xlsx'
    
    # 存工簿到响应
    wb.save(response)
    return response 

@login_required
def delete_assignment(request, assignment_id):
    if not request.user.is_teacher:
        raise PermissionDenied
        
    if request.method == 'POST':
        try:
            print(f"尝试删除作业 ID: {assignment_id}") # 调试日志
            assignment = Assignment.objects.get(id=assignment_id)
            
            # 如果文件存在，删除文件
            if assignment.file_path:
                try:
                    file_path = assignment.file_path.path
                    print(f"删除文件: {file_path}") # 调试日志
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    assignment.file_path.delete(save=False)
                except Exception as e:
                    print(f"删除文件错误: {str(e)}")
            
            # 删除相关的分发日志
            DistributionLog.objects.filter(assignment=assignment).delete()
            
            # 删除作业记录
            assignment.delete()
            print(f"作业记录删除") # 调试日志
            
            return JsonResponse({
                'status': 'success',
                'message': '作业删除成功'
            })
            
        except Assignment.DoesNotExist:
            error_msg = f'作业不存在 (ID: {assignment_id})'
            print(error_msg)
            return JsonResponse({
                'status': 'error',
                'message': error_msg
            }, status=404)
        except Exception as e:
            error_msg = f'删除作业时出错: {str(e)}'
            print(error_msg)
            return JsonResponse({
                'status': 'error',
                'message': error_msg
            }, status=500)
            
    return JsonResponse({
        'status': 'error',
        'message': '只支持POST请求'
    }, status=405)

@login_required
@require_http_methods(["POST"])
def delete_assignments(request):
    if not request.user.is_teacher:
        raise PermissionDenied
        
    try:
        data = json.loads(request.body)
        assignment_ids = data.get('assignment_ids', [])
        
        if not assignment_ids:
            return JsonResponse({
                'status': 'error',
                'message': '请选择要删除的作业'
            })
            
        assignments = Assignment.objects.filter(id__in=assignment_ids)
        deleted_count = 0
        
        for assignment in assignments:
            try:
                # 删除文件
                if assignment.file_path:
                    assignment.file_path.delete(save=False)
                # 删除记录
                assignment.delete()
                deleted_count += 1
            except Exception as e:
                print(f"删除作业 {assignment.id} 时出错: {str(e)}")
        
        return JsonResponse({
            'status': 'success',
            'message': f'成功删除 {deleted_count} 个作业任务'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': '无效的请求数据'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@login_required
def get_student_assignments(request):
    if not request.user.is_authenticated or request.user.is_teacher:
        return JsonResponse({
            'status': 'error',
            'message': '只有学生可以访问此接口'
        }, status=403)
        
    assignments = Assignment.objects.filter(
        assigned_to=request.user
    ).order_by('-upload_time')
    
    assignments_data = [{
        'id': assignment.id,
        'file_name': assignment.file_name,
        'upload_time': timezone.localtime(assignment.upload_time).strftime('%Y-%m-%d %H:%M'),
        'download_status': assignment.download_status,
        'grade': assignment.get_grade_display() if assignment.grade else None
    } for assignment in assignments]
    
    return JsonResponse({
        'status': 'success',
        'assignments': assignments_data
    })

if not os.path.exists(settings.MEDIA_ROOT):
    try:
        os.makedirs(settings.MEDIA_ROOT)
        print(f"Created MEDIA_ROOT directory: {settings.MEDIA_ROOT}")
    except Exception as e:
        print(f"创建 MEDIA_ROOT 目录失败: {str(e)}")

def test_websocket(request):
    return render(request, 'core/test_websocket.html')

def student_login(request):
    try:
        # 处理登录逻辑...
        
        # 发送WebSocket消息
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            "teacher_group",
            {
                "type": "student_status",
                "type": "student_login",
                "student": {
                    "student_id": request.user.student_id,
                    "name": request.user.first_name,
                    "ip": request.META.get('REMOTE_ADDR'),
                    "login_time": timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            }
        )
        logger.info(f"Student login WebSocket message sent for {request.user.student_id}")
        
    except Exception as e:
        logger.error(f"Error in student_login: {str(e)}", exc_info=True)
        # 处理错误...

def student_logout_view(request):
    if not request.user.is_teacher:
        student_id = request.user.student_id
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            "teacher_group",
            {
                "type": "student_status",
                "type": "student_logout",
                "student_id": student_id
            }
        )
    
    # 继续原有的登出逻辑...